import torch
import os
from PIL import Image, ImageDraw
from facenet_pytorch import MTCNN, extract_face, InceptionResnetV1
import random
import torchvision.models as models
from torchvision.models.resnet import BasicBlock
import torch.optim as optim
from torchvision import transforms
from tvqa_dataset import TVQADataset, load_json, save_json, get_test_transforms
from train import VGGFacePlus, VGGFaceSubtitle, VGGSupervised
import json
from torch.utils.data import DataLoader
from pathlib import Path
from tqdm import tqdm
import xlsxwriter
import pandas as pd
import numpy as np
from sklearn.manifold import TSNE
from sklearn.cluster import AgglomerativeClustering, KMeans, MiniBatchKMeans
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib.patheffects as PathEffects
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from matplotlib.offsetbox import OffsetImage, AnnotationBbox
import math
import torch.nn.functional as F
import umap
import umap.plot
from matplotlib.colors import LinearSegmentedColormap
from sklearn.metrics import confusion_matrix, classification_report
import logging
import argparse
from simcse import SimCSE
from train import L2Norm
from config import default_argument_parser, get_cfg, set_global_cfg, global_cfg
from fvcore.common.file_io import PathManager
from sentence_transformers import SentenceTransformer
from matplotlib.patches import Ellipse


def weighted_purity(Y, C):
    """Computes weighted purity of HAC at one particular clustering "C".
    Y, C: np.array([...]) containing unique cluster indices (need not be same!)
    Note: purity --> 1 as the number of clusters increase, so don't look at this number alone!
    """

    purity = 0.
    uniq_clid, clustering_skew = np.unique(C, return_counts=True)
    num_samples = np.zeros(uniq_clid.shape)
    # loop over all predicted clusters in C, and measure each one's cardinality and purity
    for k in uniq_clid:
        # gt labels for samples in this cluster
        k_gt = Y[np.where(C == k)[0]]
        values, counts = np.unique(k_gt, return_counts=True)
        # technically purity = max(counts) / sum(counts), but in WCP, the sum(counts) multiplies to "weight" the clusters
        purity += max(counts)

    purity /= Y.shape[0]
    return purity, clustering_skew


def NMI(Y, C):
    """Normalized Mutual Information: Clustering performance between ground-truth Y and prediction C
    Based on https://course.ccs.neu.edu/cs6140sp15/7_locality_cluster/Assignment-6/NMI.pdf
    Result matches examples on pdf
    Example:
    Y = np.array([1, 1, 1, 1, 1, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3, 3])
    C = np.array([1, 1, 1, 2, 2, 1, 1, 1, 2, 2, 2, 2, 2, 2, 2, 1, 1, 1, 1, 2])
    NMI(Y, C) = 0.1089
    C = np.array([1, 1, 1, 2, 2, 1, 1, 1, 1, 1, 1, 1, 2, 2, 2, 2, 2, 2, 2, 2])
    NMI(Y, C) = 0.2533
    """

    def entropy(labels):
        # H(Y) and H(C)
        H = 0.
        for k in np.unique(labels):
            p = (labels == k).sum() / labels.size
            H -= p * np.log2(p)
        return H

    def h_y_given_c(labels, pred):
        # H(Y | C)
        H = 0.
        for c in np.unique(pred):
            p_c = (pred == c).sum() / pred.size
            labels_c = labels[pred == c]
            for k in np.unique(labels_c):
                p = (labels_c == k).sum() / labels_c.size
                H -= p_c * p * np.log2(p)
        return H

    h_Y = entropy(Y)
    h_C = entropy(C)
    h_Y_C = h_y_given_c(Y, C)
    # I(Y; C) = H(Y) - H(Y|C)
    mi = h_Y - h_Y_C
    # NMI = 2 * MI / (H(Y) + H(C))
    nmi = 2 * mi / (h_Y + h_C)
    return nmi


def to_1D(series):
 return pd.Series([x.item() for _list in series for x in _list])


class SaveOutput:
    """
    Utility function to visualize the outputs of PCA and t-SNE
    """
    def __init__(self):
        self.outputs = []

    def __call__(self, module, module_in, module_out):
        self.outputs.append(module_out)

    def clear(self):
        self.outputs = []


def get_children(model: torch.nn.Module):
    # get children form model!
    children = list(model.children())
    flatt_children = []
    if children == []:
        # if model has no children; model is last child! :O
        return model
    else:
       # look for children from children... to the last child!
       for child in children:
            try:
                flatt_children.extend(get_children(child))
            except TypeError:
                flatt_children.append(get_children(child))
    return flatt_children


def discrete_cmap(N, base_cmap=None):
    """Create an N-bin discrete colormap from the specified input map"""

    # Note that if base_cmap is a string or None, you can simply do
    #    return plt.cm.get_cmap(base_cmap, N)
    # The following works for string, None, or a colormap instance:
    if base_cmap is None:
        return plt.cm.get_cmap(base_cmap, N)

    base = plt.cm.get_cmap(base_cmap)
    color_list = base(np.linspace(0, 1, N))
    cmap_name = base.name + str(N)
    return LinearSegmentedColormap.from_list(cmap_name, color_list, N)


def make_excel(model, dataset, source, data_path, path_to_faces, file_name):
    """
    make_excel(model=model, dataset=tvqa_all, source="dataloader",
    #            data_path="./dataset/frames_hq/friends_frames/",
    #            path_to_faces="./dataset/friends_frames/", file_name="new_all_data.xlsx")
    make_excel(model=model, dataset=tvqa_test, source="dataloader", file_name="test_data.xlsx")
    make_excel(model=model, dataset=tvqa_all, source="dataloader", file_name="all_data.xlsx")
    :param model:
    :param dataset:
    :param source:
    :param file_name:
    :return:
    """
    workbook = xlsxwriter.Workbook(f"./dataset/excel/{file_name}")
    worksheet = workbook.add_worksheet()

    if source == "dataloader":
        data_loader = DataLoader(dataset, batch_size=1, shuffle=False, num_workers=0)
        worksheet.write('A1', 'weak label')
        worksheet.write('B1', 'correct label')
        worksheet.write('C1', 'prediction')
        worksheet.write('D1', 'week=correct')
        worksheet.write('E1', 'predict=correct')
        worksheet.write('F1', 'image')
        worksheet.write('G1', 'face')
        worksheet.write('H1', 'image_loc')
        worksheet.write('I1', 'face_loc')

        for iteration, data in tqdm(enumerate(data_loader)):
            # predictions = model(data['image'])
            ann = dataset.anns[str(iteration)]
            # print(os.path.join(path_to_faces, ann["face"]))
            img = Image.open(os.path.join(path_to_faces, ann["face"]))
            draw = ImageDraw.Draw(img)
            draw.text((0, 0), f'label: {ann["name"]}', (255, 255, 255))
            # draw.text((0, 10), f'prediction: {dataset.id_to_lbs[int(predictions.argmax())]}', (255, 255, 255))
            # img.save(os.path.join(vis_path, ann["face"]))
            worksheet.write(f'A{iteration+2}', ann["name"]) #weak label
            # worksheet.write(f'C{iteration+2}', dataset.id_to_lbs[int(predictions.argmax())]) #prediction
            face = ann["face"]
            worksheet.insert_image(f'F{iteration+2}', os.path.join(data_path, face[0:28]+f"/{face.split('_')[-2]}.jpg")) #image
            # face: friends_s01e01_seg02_clip_03_00108_0.png
            # tvqa_experiment/dataset/frames_hq/friends_frames/friends_s01e01_seg02_clip_03/00108.jpg
            worksheet.insert_image(f'G{iteration + 2}', os.path.join(path_to_faces, ann["face"]))  # face
            worksheet.write(f'H{iteration + 2}', face[0:28]+f"/{face.split('_')[-2]}.jpg") # image_loc
            worksheet.write(f'I{iteration + 2}', ann["face"]) # face_loc

        workbook.close()


def make_excel_for_test(new_anns):
    import xlsxwriter
    data_path = "./dataset/frames_hq/friends_frames/"
    path_to_faces = "./dataset/friends_frames/"
    workbook = xlsxwriter.Workbook(f"./dataset/excel/s01e01_new_new.xlsx")
    worksheet = workbook.add_worksheet()
    worksheet.write('A1', 'weak label')
    worksheet.write('B1', 'correct label')
    worksheet.write('C1', 'prediction')
    worksheet.write('D1', 'week=correct')
    worksheet.write('E1', 'predict=correct')
    worksheet.write('F1', 'image')
    worksheet.write('G1', 'face')
    worksheet.write('H1', 'image_loc')
    worksheet.write('I1', 'face_loc')
    old_anns = load_json("./dataset/all_annotations.json")
    for iteration, ann in enumerate(new_anns.values()):
        worksheet.write(f'A{iteration + 2}', "".join(ann["name"]))  # weak label
        face = ann["face"]
        worksheet.insert_image(f'F{iteration + 2}',
                               os.path.join(data_path, face[0:28] + f"/{face.split('_')[-2]}.jpg"))
        worksheet.insert_image(f'G{iteration + 2}', os.path.join(path_to_faces, ann["face"]))  # face
        worksheet.write(f'H{iteration + 2}', face[0:28] + f"/{face.split('_')[-2]}.jpg")  # image_loc
        worksheet.write(f'I{iteration + 2}', ann["face"])  # face_loc
        for old_ann in old_anns.values():
            if ann["face"] == old_ann["face"]:
                worksheet.write(f'B{iteration + 2}', old_ann["target_name"])  # face_loc
    workbook.close()


def visualize_embeddings(model, dataset, file_name, method="tsne", model_mode="facenet_pretrained", normalize=False, mode="nothing", preload=False):
    """
    # visualize_embeddings(model=model, dataset=tvqa_all, file_name="all_data", method="tsne", model_mode=model_mode, normalize=False, mode="nothing", preload=True)
    build_embeddings(model=model, dataset=tvqa_all, exp=exp, file_name="all_data", method="umap")
    :param model: pretrained model that generates face embeddings
    :param dataset: pointing to the data : tvqa_all, tvqa_train, tvqa_test, tvqa_val
    :param file_name: excel file name -> all_data
    :param method: tsne or umap
    :param model_mode: facenet_pretrained or resnet_pretrained
    :param normalize: whether to normalize embeddings or not
    :param mode: pictures, text or nothing
    :return: saves files of such visualizations
    """
    dataset_loc = "./dataset/excel"

    embeddings = calculate_embeddings(model, dataset, model_mode, normalize=normalize, preload=preload)
    if method == "tsne":
        # dists = [[(e1 - e2).norm().item() for e2 in embeddings] for e1 in embeddings]
        # print(pd.DataFrame(dists)) #, columns=names, index=names??
        print("TSNE is being calculated...")
        tsne_grid = TSNE(random_state=10, n_iter=2000).fit_transform(embeddings.detach().numpy())
        print("TSNE is calculated!")
        df = pd.read_excel(f"{dataset_loc}/{file_name}.xls", usecols="A,B,C")  # weak label, correct label, prediction
        workbook = openpyxl.load_workbook(f"{dataset_loc}/{file_name}.xlsx")
        sheet = workbook['Sheet1']
        image_loader = SheetImageLoader(sheet)
        hac8 = AgglomerativeClustering(n_clusters=8).fit_predict(embeddings.detach().numpy())

        dataset.lbl_to_id['Unknown'] = 50
        dataset.id_to_lbs[50] ='Unknown'

        # hac8 = [dataset.id_to_lbs[predicted_lbl] for predicted_lbl in hac8]
        fig, plt = visualize_tsne(tsne_grid, hac8, dataset.id_to_lbs, image_loader, mode=mode)
        fig.savefig(os.path.join("", f"hac8_clusteringpredictions_{file_name}_{mode}pictures_{model_mode}.pdf"))
        plt.clf()

        correct_ids = [dataset.lbl_to_id[lbl] for lbl in df['correct label'][:-1].values.tolist()]
        fig, plt = visualize_tsne(tsne_grid, correct_ids, dataset.id_to_lbs, image_loader, mode=mode)
        fig.savefig(os.path.join("", f"tsne_corrects_{file_name}_{mode}pictures_{model_mode}.pdf"))
        plt.clf()

        prediction_ids = [dataset.lbl_to_id[lbl] for lbl in df['prediction'][:-1].values.tolist()]
        fig, plt = visualize_tsne(tsne_grid, prediction_ids, dataset.id_to_lbs, image_loader, mode=mode)
        fig.savefig(os.path.join("", f"tsne_predictions_{file_name}_{mode}pictures_{model_mode}.pdf"))
        plt.clf()

    elif method == "umap":
        import matplotlib.pyplot as plt
        df = pd.read_excel(f"{dataset_loc}/{file_name}.xls", usecols="A,B,C")
        mapper = umap.UMAP().fit(embeddings.detach().numpy())
        colors = df['correct label'][:-1]
        targets = np.asarray([lbl for lbl in colors])
        fig, ax = plt.subplots()
        umap.plot.points(mapper, labels=targets, ax=ax)
        fig.savefig(f"umap_{file_name}_{model_mode}.pdf")


def calculate_embeddings(model, dataset, emb_path, model_mode="facenet_pretrained", normalize=False, preload=False):
    # emb_path = "./dataset/embeddings.pt"
    if preload is True:
        print("Preloading from existing embedding.pt file!")
        return torch.load(emb_path)
    print("Calculating the embeddings and saving them in embedding.pt file!")
    # data_loader = DataLoader(dataset, batch_size=len(dataset), shuffle=False, num_workers=0)  # todo batch size
    data_loader = DataLoader(dataset, batch_size=4096, shuffle=False, num_workers=0)
    flattened_model = get_children(model)
    if model_mode == "resnet_pretrained" or model_mode == "facenet_reclassified":
        save_output = SaveOutput()
        hook_handles = []
        for layer in flattened_model:
            handle = layer.register_forward_hook(save_output)
            hook_handles.append(handle)
        with torch.no_grad():
            for iteration, data in tqdm(enumerate(data_loader)):
                model(data[0]['image'][None, :, :, :])

                if model_mode == "resnet_pretrained":
                    temp_emb = save_output.outputs[len(save_output.outputs) - 2]  # embeddings from the layer before logits
                elif model_mode == "facenet_reclassified":
                    temp_emb = save_output.outputs[-1]  # embeddings from the layer before loss layer
                if iteration == 0:
                    embeddings = temp_emb
                else:
                    embeddings = torch.cat((embeddings, temp_emb), 0)
                save_output.clear()


    elif model_mode == "facenet_pretrained":
        with torch.no_grad():
            for iteration, data in tqdm(enumerate(data_loader)):
                if iteration == 0:
                    # embeddings = model(data[0]['image'][None, :, :, :])
                    embeddings = model(data['image'][:, :, :])
                else:
                    # embeddings = torch.cat((embeddings, model(data[0]['image'][None, :, :, :])), 0)
                    embeddings = torch.cat((embeddings, model(data['image'][:, :, :])), 0)

    if normalize:
        embeddings = F.normalize(embeddings, p=2, dim=1)

    torch.save(embeddings, emb_path)
    return embeddings


def calculate_sentence_embeddings(dataset, emb_path, normalize=False, preload=False):
    if preload is True:
        print(f"Preloading from existing {emb_path} file!")
        return torch.load(emb_path)
    print("Calculating the embeddings and saving them in sentence_embedding.pt file!")
    data_loader = DataLoader(dataset, batch_size=1, shuffle=False, num_workers=0, collate_fn=lambda x: x)
    model = SentenceTransformer('sentence-transformers/paraphrase-mpnet-base-v2')
    # model = SimCSE("princeton-nlp/sup-simcse-bert-base-uncased")

    with torch.no_grad():
        for iteration, data in tqdm(enumerate(data_loader)):
            # weak_lbls = data[0]['name']

            sub_emb = torch.from_numpy(model.encode(data[0]['subtitle'])[None, :])
            # if weak_lbls:
            #     weak_lbls_embedding = model.encode(' '.join(weak_lbls))[None, :]
            # else:
            #     weak_lbls_embedding = sub_emb
            #
            # sub_emb = torch.cat((sub_emb, weak_lbls_embedding), 1)

            if iteration == 0:
                all_embeddings = sub_emb
            else:
                all_embeddings = torch.cat((all_embeddings, sub_emb), 0)

    if normalize:
        all_embeddings = F.normalize(all_embeddings, p=2, dim=1)

    torch.save(all_embeddings, emb_path)
    return all_embeddings


def calculate_joint_embeddings(model, dataset, emb_path, model_mode="facenet_pretrained", normalize=False, preload=False):
    # emb_path = "./dataset/embeddings.pt"
    if preload is True:
        print("Preloading from existing embedding.pt file!")
        return torch.load(emb_path)
    print("Calculating the embeddings and saving them in embedding.pt file!")
    # data_loader = DataLoader(dataset, batch_size=len(dataset), shuffle=False, num_workers=0)  # todo batch size
    data_loader = DataLoader(dataset, batch_size=1, shuffle=False, num_workers=0, collate_fn=lambda x: x)
    flattened_model = get_children(model)

    save_output = SaveOutput()
    hook_handles = []
    for layer in flattened_model:
        handle = layer.register_forward_hook(save_output)
        hook_handles.append(handle)
    with torch.no_grad():
        for iteration, data in tqdm(enumerate(data_loader)):
            model(data[0]['image'][None, :, :, :], None, data[0]['subtitle'], None)

            temp_emb = save_output.outputs[-1]  # embeddings from the layer before loss layer
            if iteration == 0:
                embeddings = temp_emb
            else:
                embeddings = torch.cat((embeddings, temp_emb), 0)
            save_output.clear()

    if normalize:
        embeddings = F.normalize(embeddings, p=2, dim=1)

    torch.save(embeddings, emb_path)
    return embeddings


#todo: https://jakevdp.github.io/PythonDataScienceHandbook/05.12-gaussian-mixtures.html
def draw_ellipse(position, covariance, ax=None, **kwargs):
    """Draw an ellipse with a given position and covariance"""
    # ax = ax or plt.gca()

    # Convert covariance to principal axes
    if covariance.shape == (2, 2):
        U, s, Vt = np.linalg.svd(covariance)
        angle = np.degrees(np.arctan2(U[1, 0], U[0, 0]))
        width, height = 2 * np.sqrt(s)
    else:
        angle = 0
        # width, height = 2 * np.sqrt(covariance)
        width = 2 * np.sqrt(covariance)
        height = width

    # Draw the Ellipse
    # for w, h in zip(width, height):
    #     ax.add_patch(Ellipse(position, w, h, angle, **kwargs))
    for nsig in range(1, 4):
        ax.add_patch(Ellipse(position, nsig * width, nsig * height, angle, **kwargs))


def plot_gmm(gmm, labels, tsne_grid, label=True):
    fig = plt.figure(figsize=(8, 8))
    ax = plt.subplot(aspect='equal')

    # ax = ax or plt.gca()
    # labels = gmm.fit(X).predict(X)
    # if label:
    #     ax.scatter(tsne_grid[:, 0], tsne_grid[:, 1], c=labels, s=40, cmap='viridis', zorder=2)
    # else:
    #     ax.scatter(tsne_grid[:, 0], tsne_grid[:, 1], s=40, zorder=2)
    # ax.axis('equal')

    w_factor = 0.2 / gmm.weights_.max()
    for pos, covar, w in zip(gmm.means_, gmm.covariances_, gmm.weights_):
        draw_ellipse(pos, covar, ax=ax, alpha=w * w_factor)

    return fig, plt


def visualize_tsne(tsne_grid, label_ids, id_to_lbl, image_loader=None, mode="nothing"):

    num_classes = len(id_to_lbl)
    # convert to pandas
    label_ids = pd.DataFrame(label_ids, columns=['label'])['label']
    # create a scatter plot.
    fig = plt.figure(figsize=(8, 8))
    ax = plt.subplot(aspect='equal')
    if not label_ids.isnull().values.any():
        plt.scatter(tsne_grid[:, 0], tsne_grid[:, 1], lw=0, s=40, c=np.asarray(label_ids),
                    cmap=discrete_cmap(num_classes, "tab10"))
        # , c = palette[np.asarray([lbl_to_id[lbl] for lbl in colors])]
        # c = np.random.randint(num_classes, size=len(tsne_grid[:, 1]))
    else:
        plt.scatter(tsne_grid[:, 0], tsne_grid[:, 1], lw=0, s=40)
    plt.xlim(-25, 25)
    plt.ylim(-25, 25)
    cbar = plt.colorbar(ticks=range(num_classes))
    cbar.set_ticklabels(list(id_to_lbl.values()))
    plt.clim(-0.5, num_classes - 0.5)
    ax.axis('off')
    ax.axis('tight')

    if mode == "picture":
        max_dim = 16
        for i, (x, y) in enumerate(tsne_grid):
            print(i, x, y)
            tile = image_loader.get(f'G{i+2}')
            rs = max(1, tile.width/max_dim, tile.height/max_dim)
            tile = tile.resize((int(tile.width/rs), int(tile.height/rs)), Image.ANTIALIAS)
            imagebox = OffsetImage(tile) #, zoom=0.2)
            ab = AnnotationBbox(imagebox, (x, y), pad=0.1)
            ax.add_artist(ab)

    if mode == "text":
        # add the labels for each digit corresponding to the label
        if not label_ids.isnull().values.any():
            txts = []
            for id, lbl in id_to_lbl.items():
                # Position of each label at median of data points.
                xtext, ytext = np.median(tsne_grid[np.asarray(label_ids) == id, :], axis=0)
                if math.isnan(xtext) or math.isnan(ytext):  # this label does not exist in this set
                    continue
                txt = ax.text(xtext, ytext, lbl, fontsize=10, zorder=100)
                txt.set_path_effects([
                    PathEffects.Stroke(linewidth=2, foreground="w"),
                    PathEffects.Normal()])
                txts.append(txt)

    return fig, plt


def evaluate(model, dataset, model_mode):
    data_loader = DataLoader(dataset, batch_size=1, shuffle=False, num_workers=0)

    with torch.no_grad():
        correct_target_names = []
        correct_target_ids = []
        target_ids = []
        predictions = []
        for data in tqdm(data_loader):
            # target_ids.append(data['weak_label'])
            correct_target_names.extend(data['correct_target_name'])
            correct_target_ids.append(data['correct_target_id'])
            if model_mode == "facenet_pretrained":
                predictions = np.nan
                continue
            else:
                prediction = model(data['image'])
                predictions.append(int(prediction.argmax()))

    # flatten list of lists
    # target_ids = [item.item() for sublist in target_ids for item in sublist]
    correct_target_ids = [item.item() for sublist in correct_target_ids for item in sublist]

    results = pd.DataFrame({'correct_target_name': correct_target_names,
                            'correct_target_id': correct_target_ids,
                            'model_prediction': predictions,
                            # 'weak_label': target_ids,
                            # 'max_cluster_prediction': np.nan,
                            })

    return results


def clustering_with_gmm(results, face_embeddings, tsne_grid, n_clusters=7):
    from sklearn import mixture
    # results["gmm"] = np.nan
    model = mixture.GaussianMixture(n_components=n_clusters, covariance_type='spherical')
    # a = model.fit(face_embeddings)
    # labels = a.predict(face_embeddings)
    labels = model.fit(tsne_grid).predict(tsne_grid)
    results["gmm"] = labels
    # probs = model.predict_proba(face_embeddings)
    plt, fig = plot_gmm(model, labels, tsne_grid)
    return results, plt, fig


def build_graph(embeddings, train_dataset, neighbours=3385, no_edge=True):
    from sklearn.neighbors import kneighbors_graph
    train_loader = DataLoader(train_dataset, batch_size=1, shuffle=False, num_workers=0)
    knn_graph = kneighbors_graph(embeddings, neighbours, include_self=True)  # make the full neighbour graph
    knn_graph = knn_graph.toarray()

    if no_edge:
        images = []
        with torch.no_grad():
            for i, data in enumerate(train_loader):
                img = '_'.join(data["face"][0].split("_")[:-1])
                images.append(img)

        for i1, img1 in enumerate(images):
            for i2, img2 in enumerate(images):
                if img1 == img2 and i1 != i2:
                    knn_graph[i1][i2] = 0.0

    return knn_graph


def predict_with_clustering(results, embeddings, n_clusters, knn_graph=None, pred_mode="max_cluster_prediction"):
    # from sklearn import mixture
    # model = mixture.GaussianMixture(n_components=n_clusters, covariance_type='spherical')
    # labels = model.fit(embeddings).predict(embeddings)
    # results["gmm"] = labels
    # results[pred_mode] = np.nan
    embeddings = embeddings.detach().numpy() #.astype('float32')
    # print(embeddings.dtype)
    # logger.info(f"type of embeddings is {embeddings.dtype}")
    if global_cfg.TRAINING.clustering == "AgglomerativeClustering":
        hac8_id = AgglomerativeClustering(n_clusters=n_clusters).fit_predict(embeddings)
    elif global_cfg.TRAINING.clustering == "KMeans":
        hac8_id = KMeans(n_clusters=n_clusters).fit_predict(embeddings)
    elif global_cfg.TRAINING.clustering == "MiniBatchKMeans":
        hac8_id = MiniBatchKMeans(batch_size=global_cfg.TRAINING.kmeans_batch_size, n_clusters=n_clusters).fit_predict(embeddings)

    # import pdb
    # pdb.set_trace()
    # if knn_graph is None:
    #     hac8_id = AgglomerativeClustering(n_clusters=n_clusters).fit_predict(embeddings)
    # else:
    #     hac8_id = AgglomerativeClustering(n_clusters=n_clusters, connectivity=knn_graph).fit_predict(embeddings.detach().numpy())
    hac8_id = pd.DataFrame(hac8_id, columns=['label'])

    for i in range(n_clusters):
        # true for all indices equal to that cluster
        # if np.all(results.loc[hac8_id['label'] == 18, "weak_label_ids"]) == []: #todo: what have i done here??
        #     results.loc[hac8_id['label'] == i, pred_mode] = 0.6 #unknown
        # else:
        results.loc[hac8_id['label'] == i, pred_mode] = to_1D(results.loc[hac8_id['label'] == i, "weak_label_ids"]).value_counts().idxmax()

    results[pred_mode] = pd.to_numeric(results[pred_mode], downcast='integer')

    results['direct'] = results['weak_label_ids']
    results['direct'] = results['direct'].apply(lambda x: x[0].item() if len(x) == 1 else np.nan)
    results['M2'] = results[pred_mode]
    results.loc[results['direct'].notnull(), 'M2'] = results.loc[results['direct'].notnull(), 'direct']
    results['M2'] = pd.to_numeric(results['M2'], downcast='integer')

    clustering_ids = hac8_id

    return results, clustering_ids


def calc_accuracies(results, mode="max_cluster_prediction"):
    correct = (results[mode] == results["correct_target_id"]).value_counts().loc[True]
    incorrect = (results[mode] == results["correct_target_id"]).value_counts().loc[False]
    accuracy = correct / (correct + incorrect)
    # from sklearn.metrics import classification_report, accuracy_score
    # or accuracy_score(results['correct_target_id'], results['max_cluster_prediction'])
    return accuracy


def calc_accuracies_bbt(gt_id, predictions_id):
    correct = np.count_nonzero(predictions_id == gt_id)
    incorrect = len(predictions_id) - np.count_nonzero(predictions_id == gt_id)
    accuracy = correct / (correct + incorrect)
    # from sklearn.metrics import classification_report, accuracy_score
    # or accuracy_score(results['correct_target_id'], results['max_cluster_prediction'])
    return accuracy


def calc_per_class_prec_recall(results, mode="max_cluster_prediction"):
    return classification_report(results['correct_target_id'], results[mode], digits=3)


def calc_per_class_prec_recall_bbt(train_dataset, gt_id, predictions_id):
    return classification_report(gt_id, predictions_id, labels=list(train_dataset.lbl_to_id.keys()), digits=3)


def calc_per_class_accuracy(dataset, results, mode="max_cluster_prediction"):
    cm = confusion_matrix(results['correct_target_id'], results[mode])
    num_classes = len(dataset.lbl_to_id.keys())
    plt.imshow(cm, cmap='plasma', interpolation='nearest')

    plt.xticks(range(num_classes), np.array(list(dataset.lbl_to_id.keys())), rotation=90, fontsize=6)
    plt.yticks(range(num_classes), np.array(list(dataset.lbl_to_id.keys())), rotation=0, fontsize=6)

    # Plot a colorbar with label.
    cb = plt.colorbar()
    cb.set_label("Number of predictions")

    # Add title and labels to plot.
    plt.title("Confusion Matrix for predictions and correct labels")
    plt.xlabel('Correct Label')
    plt.ylabel('Predicted Label')
    plt.savefig('confusion_matrix_upperbound.pdf')
    plt.clf()


    cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
    return cm.diagonal(), dataset.id_to_lbs


def calc_per_class_accuracy_bbt(dataset, gt_names, predictions_names):
    cm = confusion_matrix(gt_names, predictions_names, labels=list(dataset.lbl_to_id.keys()))
    num_classes = len(dataset.lbl_to_id.keys())
    plt.imshow(cm, cmap='plasma', interpolation='nearest')

    plt.xticks(range(num_classes), np.array(list(dataset.lbl_to_id.keys())), rotation=90, fontsize=6)
    plt.yticks(range(num_classes), np.array(list(dataset.lbl_to_id.keys())), rotation=0, fontsize=6)

    # Plot a colorbar with label.
    cb = plt.colorbar()
    cb.set_label("Number of predictions")

    # Add title and labels to plot.
    plt.title("Confusion Matrix for predictions and correct labels")
    plt.xlabel('Correct Label')
    plt.ylabel('Predicted Label')
    plt.savefig('confusion_matrix_upperbound.pdf')
    plt.clf()


    cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
    return cm.diagonal(), dataset.id_to_lbs


def exp_num_clusters(cluster_range, results, embeddings, test_dataset):
    accuracies_per_cluster = []
    clusters = []
    for n in cluster_range:
        results, clustering_ids = predict_with_clustering(results, embeddings, n_clusters=n)
        prediction_mode = "max_cluster_prediction"

        print(
            f"per sample accuracy is {calc_accuracies(results, mode=prediction_mode)}")
        accuracies = calc_per_class_accuracy(test_dataset, results, mode=prediction_mode)
        print(f"mean per class accuracy: {accuracies[0].mean()}")
        print(f"per class accuracies: {accuracies}")
        accuracies_per_cluster.append(accuracies[0].mean())
        print(f"per class precision and recalls: {calc_per_class_prec_recall(results, mode=prediction_mode)}")

        clusters.append(n)

    fig = plt.figure(figsize=(8, 6))
    ax = plt.subplot(aspect='equal')
    # plt.style.use('seaborn-darkgrid')
    # plt.yticks(np.arange(0.0, 1.1, 0.1))
    ax.axis('tight')
    plt.ylim(0.0, 1.0)
    plt.xlim(6, 31)
    plt.bar(x=clusters, height=accuracies_per_cluster, width=0.4, color='#c3abd0')  # width
    plt.plot(clusters, accuracies_per_cluster, color='#815f76')
    plt.ylabel("Accuracy", rotation=90)
    plt.xlabel("Number of Clusters")
    fig.savefig(os.path.join(f"./output/ablation_num_clusters_friends/number_of_clusters_2.pdf"))
    plt.clf()


def cleanse_labels(dataset, results, file_path):
    anns = dataset.anns
    for ann, max_prediction in zip(anns.values(), results['max_cluster_prediction']):
        ann['cleansed'] = dataset.id_to_lbs[max_prediction]
    save_json(anns, file_path=file_path)


def save_predictions(id_to_lbs, results, file_path, prediction_dict, prediction_mode='cleansed'):
    for ann, max_prediction in zip(prediction_dict.values(), results[prediction_mode]):
        ann[prediction_mode] = id_to_lbs[max_prediction]
    save_json(prediction_dict, file_path=file_path)
    return prediction_dict


def evaluate_self_supervised(train_dataset, test_dataset, face_embeddings, mode="baseline1"):
    train_loader = DataLoader(train_dataset, batch_size=1, shuffle=False, num_workers=0)
    test_loader = DataLoader(test_dataset, batch_size=1, shuffle=False, num_workers=0)

    with torch.no_grad():
        correct_target_names = []
        correct_target_ids = []
        weak_ids = []
        cleansed_ids = []
        face_emb = []
        unknown_id = torch.tensor([train_dataset.lbl_to_id["Unknown"]])
        for train_data, test_data, face_embedding in tqdm(zip(train_loader, test_loader, face_embeddings)):
            if mode == "baseline0":
                if train_data['weak_id']:
                    choice = random.choice(train_data['weak_id'])
                    weak_ids.append(choice if choice in list(train_dataset.id_to_lbs.keys()) else unknown_id)
                else:
                    weak_ids.append(unknown_id)
            elif mode == "baseline1":
                weak_ids.append(train_data['weak_id'])
            correct_target_names.extend(test_data['correct_target_name'])
            correct_target_ids.append(test_data['correct_target_id'])
            cleansed_ids.append(train_dataset.lbl_to_id[train_data['cleansed'][0]])
            face_emb.append(face_embedding)

    # flatten list of lists
    # weak_ids = [item.item() for sublist in weak_ids for item in sublist]
    correct_target_ids = [item.item() for sublist in correct_target_ids for item in sublist]

    if mode == "baseline0":
        results = pd.DataFrame({'correct_target_name': correct_target_names,
                                'correct_target_id': correct_target_ids,
                                'random_weak_label': weak_ids,
                                })
    elif mode == "baseline1":
        results = pd.DataFrame({'correct_target_name': correct_target_names,
                                'correct_target_id': correct_target_ids,
                                'weak_label_ids': weak_ids,
                                'max_cluster_prediction': np.nan,
                                'cleansed': cleansed_ids,
                                'subtitle_prediction': np.nan,
                                'face_embedding': face_emb,
                                '0': np.nan,
                                '1': np.nan,
                                '2': np.nan,
                                '3': np.nan,
                                '4': np.nan,
                                '5': np.nan,
                                '6': np.nan,
                                'min_distance': np.nan,
                                'closest_cluster': np.nan,
                                })

    return results


def prepare_result(train_dataset, face_embeddings):
    train_loader = DataLoader(train_dataset, batch_size=1, shuffle=False, num_workers=0)

    with torch.no_grad():
        weak_ids = []
        cleansed_ids = []
        face_emb = []
        unknown_id = torch.tensor([train_dataset.lbl_to_id["Unknown"]])
        for train_data, face_embedding in tqdm(zip(train_loader, face_embeddings)):
            weak_ids.append(train_data['weak_id'])
            cleansed_ids.append(train_dataset.lbl_to_id[train_data['cleansed'][0]])
            face_emb.append(face_embedding)

    # flatten list of lists
    # correct_target_ids = [item.item() for sublist in correct_target_ids for item in sublist]

    results = pd.DataFrame({'face_embedding': face_emb,
                            'weak_label_ids': weak_ids,
                            'correct_target_name': np.nan,
                            'correct_target_id': np.nan,
                            'max_cluster_prediction': np.nan,
                            'cleansed': cleansed_ids,
                            '0': np.nan,
                            '1': np.nan,
                            '2': np.nan,
                            '3': np.nan,
                            '4': np.nan,
                            '5': np.nan,
                            '6': np.nan,
                            '7': np.nan,
                            '8': np.nan,
                            'min_distance': np.nan,
                            'closest_cluster': np.nan,
                            })

    return results


def evaluate_ep_1_5(train_dataset, face_embeddings, mode="baseline1"):
    train_loader = DataLoader(train_dataset, batch_size=1, shuffle=False, num_workers=0)
    # test_loader = DataLoader(test_dataset, batch_size=1, shuffle=False, num_workers=0)

    with torch.no_grad():
        correct_target_names = []
        correct_target_ids = []
        weak_ids = []
        cleansed_ids = []
        face_emb = []
        unknown_id = torch.tensor([train_dataset.lbl_to_id["Unknown"]])
        for train_data, face_embedding in tqdm(zip(train_loader, face_embeddings)):
            if mode == "baseline0":
                if train_data['weak_id']:
                    choice = random.choice(train_data['weak_id'])
                    weak_ids.append(choice if choice in list(train_dataset.id_to_lbs.keys()) else unknown_id)
                else:
                    weak_ids.append(unknown_id)
            elif mode == "baseline1":
                weak_ids.append(train_data['weak_id'])
            correct_target_names.extend(train_data['correct_target_name'])
            correct_target_ids.append(train_data['correct_target_id'])
            cleansed_ids.append(train_dataset.lbl_to_id[train_data['cleansed'][0]])
            face_emb.append(face_embedding)

    # flatten list of lists
    # weak_ids = [item.item() for sublist in weak_ids for item in sublist]
    correct_target_ids = [item.item() for sublist in correct_target_ids for item in sublist]

    if mode == "baseline0":
        results = pd.DataFrame({'correct_target_name': correct_target_names,
                                'correct_target_id': correct_target_ids,
                                'random_weak_label': weak_ids,
                                })
    elif mode == "baseline1":
        results = pd.DataFrame({'correct_target_name': correct_target_names,
                                'correct_target_id': correct_target_ids,
                                'weak_label_ids': weak_ids,
                                'max_cluster_prediction': np.nan,
                                'cleansed': cleansed_ids,
                                'face_embedding': face_emb,
                                '0': np.nan,
                                '1': np.nan,
                                '2': np.nan,
                                '3': np.nan,
                                '4': np.nan,
                                '5': np.nan,
                                '6': np.nan,
                                'min_distance': np.nan,
                                'closest_cluster': np.nan,
                                })

    return results


def evaluate_oracle_supervised(model, test_dataset, face_embeddings, mode="baseline1"):
    # train_loader = DataLoader(train_dataset, batch_size=1, shuffle=False, num_workers=0)
    test_loader = DataLoader(test_dataset, batch_size=1, shuffle=False, num_workers=0)

    with torch.no_grad():
        correct_target_names = []
        correct_target_ids = []
        weak_ids = []
        cleansed_ids = []
        face_emb = []
        predictions = []
        unknown_id = torch.tensor([test_dataset.lbl_to_id["Unknown"]])
        for test_data, face_embedding in tqdm(zip(test_loader, face_embeddings)):
            if mode == "baseline0":
                if test_data['weak_label']:
                    choice = random.choice(test_data['weak_id'])
                    weak_ids.append(choice if choice in list(test_dataset.id_to_lbs.keys()) else unknown_id)
                else:
                    weak_ids.append(unknown_id)
            elif mode == "baseline1":
                weak_ids.append(test_data['weak_id'])
            correct_target_names.extend(test_data['correct_target_name'])
            correct_target_ids.append(test_data['correct_target_id'])
            cleansed_ids.append(test_dataset.lbl_to_id[test_data['cleansed'][0]])
            face_emb.append(face_embedding)
            prediction = model(test_data['image'])
            predictions.append(int(prediction.argmax()))

    # flatten list of lists
    # weak_ids = [item.item() for sublist in weak_ids for item in sublist]
    correct_target_ids = [item.item() for sublist in correct_target_ids for item in sublist]

    if mode == "baseline0":
        results = pd.DataFrame({'correct_target_name': correct_target_names,
                                'correct_target_id': correct_target_ids,
                                'random_weak_label': weak_ids,
                                })
    elif mode == "baseline1":
        results = pd.DataFrame({
            'correct_target_name': correct_target_names,
                                'correct_target_id': correct_target_ids,
                                'weak_label_ids': weak_ids,
                                'max_cluster_prediction': np.nan,
                                'cleansed': cleansed_ids,
                                'face_embedding': face_emb,
                                'model_prediction': predictions,
                                '0': np.nan,
                                '1': np.nan,
                                '2': np.nan,
                                '3': np.nan,
                                '4': np.nan,
                                '5': np.nan,
                                '6': np.nan,
                                'min_distance': np.nan,
                                'closest_cluster': np.nan,
                                })

    return results


def match_bbox(train_dataset, test_dataset):
    matchings = {str(i): {"matching": [], "iou": []} for i in range(len(test_dataset))}
    for i_test, s_test in tqdm(test_dataset.anns.items()):
        for i_train, s_train in train_dataset.anns.items():
            if s_test['clip'] + '_' + s_test['img'] == s_train['clip'] + '_' + s_train['img']:
                iou = calc_iou(s_test['bbox'], s_train['bbox'])
                # if iou > 0.0: #0.7
                if matchings[i_test]['matching']:
                    # print("more matchings!")
                    matchings[i_test]['matching'].append(i_train)
                    matchings[i_test]['iou'].append(iou)
                else:
                    matchings[i_test]['matching'] = [i_train]
                    matchings[i_test]['iou'] = [iou]
    save_json(matchings, "./dataset/box_matchings_iou0.json")
    return matchings


def match_bbox_2(test_dataset, prediction_dict):

    ########################calculate matchings########################
    if os.path.exists("./dataset/box_matchings_with_iou.json"):
        print("Found matches json, loading ...")
        return load_json("./dataset/box_matchings_with_iou.json")

    else:
        print("No matches.json, creating ...")
        matchings = {str(i): {"matching": [], "iou": [], "prediction_baseline_0": []} for i in range(len(test_dataset.anns))}
        for i_test, s_test in tqdm(test_dataset.anns.items()):
            for train_face, train_ann in prediction_dict.items():
                if s_test['clip'] + '_' + s_test['img'] == train_ann['clip'] + '_' + train_ann['img']:
                    iou = calc_iou(s_test['bbox'], train_ann['bbox'])
                    # more than one match
                    if matchings[i_test]['matching']:
                        matchings[i_test]['matching'].append(train_face)
                        matchings[i_test]['iou'].append(iou)
                        matchings[i_test]['prediction_baseline_0'].append(train_ann['prediction_baseline_0'])
                    else:
                        matchings[i_test]['matching'] = [train_face] #matched faces
                        matchings[i_test]['iou'] = [iou] #iou of the matched faces
                        matchings[i_test]['prediction_baseline_0'] = [train_ann['prediction_baseline_0']] #predictions of the matched faces
        save_json(matchings, "./dataset/box_matchings.json")


    #add here the iou code:
    to_be_deleted_ind = []
    to_be_deleted_face = []
    for i_test, s_test in tqdm(test_dataset.anns.items()):
        if not matchings[i_test]["iou"]:
            # this face is not matched with anything -> discard
            to_be_deleted_ind.append(i_test)
            to_be_deleted_face.append(s_test['face'])
        else:
            max_iou = np.max(np.array(matchings[i_test]["iou"]))
            if max_iou == 0:
                # this face is not matched with anything -> discard
                to_be_deleted_ind.append(i_test)
                to_be_deleted_face.append(s_test['face'])
    save_json(to_be_deleted_ind, "./dataset/to_be_deleted_test_index.json")
    save_json(to_be_deleted_face, "./dataset/to_be_deleted_test_face.json")
    # update matchings:
    valid_test_anns = {}
    to_be_deleted_test_ind = load_json("./dataset/to_be_deleted_test_index.json")
    j = 0
    for i_test, ann_test in tqdm(test_dataset.anns.items()):
        if i_test not in to_be_deleted_test_ind:
            valid_test_anns[str(j)] = ann_test
            j += 1

    save_json(valid_test_anns, "./dataset/bbt_test_annotations.json")

    new_matchings = {}
    to_be_deleted_test_ind = load_json("./dataset/to_be_deleted_test_index.json")
    j = 0
    for i_match, match in tqdm(matchings.items()):
        if i_match not in to_be_deleted_test_ind:
            new_matchings[str(j)] = match
            j += 1
    save_json(new_matchings, "./dataset/box_matchings_with_iou.json")
    return new_matchings


def calculate_bbt(matchings, prediction_dict, test_dataset, prediction_mode='closest_cluster'):
    # train_dataset is basically prediction_dict, it is the file that is exactly like train annotations but with preedictions
    # prediction_dict = train_dataset.anns
    face_to_train_idx = load_json("./dataset/face_to_train_idx.json")
    test_9_10_idx_to_test_all_idx = load_json("./dataset/test_9_10_idx_to_test_all_idx.json")

    gt_names = []
    gt_ids = []
    predictions = []
    for i_test, s_test in tqdm(test_dataset.anns.items()):
        # if s_test['face'].startswith(('s09', 's10')): #todo
        gt_names.append(s_test['name'])
        gt_ids.append(test_dataset.lbl_to_id[s_test['name']])
        if global_cfg.TRAINING.exp_type == "oracle" and global_cfg.TRAINING.ours_or_baseline == "ours":
            new_test_idx = test_9_10_idx_to_test_all_idx[i_test]
            i_test = new_test_idx

        if global_cfg.TRAINING.ours_or_baseline == "baseline":
            predictions.append(prediction_dict[i_test][prediction_mode])
        elif global_cfg.TRAINING.ours_or_baseline == "ours":
            iou_idx = np.array(matchings[i_test]["iou"]).argmax()
            matched_face = matchings[i_test]['matching'][iou_idx]
            predictions.append(prediction_dict[face_to_train_idx[matched_face]][prediction_mode])


    # gt_names = []
    # gt_ids = []
    # predictions = []
    # for i_test, s_test in tqdm(test_dataset.anns.items()):
    #     if matchings[i_test]["iou"]:
    #         if np.max(np.array(matchings[i_test]["iou"])) != 0:
    #             gt_names.append(s_test['name'])
    #             gt_ids.append(test_dataset.lbl_to_id[s_test['name']])
    #             predictions.append('None')
    #
    #             iou_idx = np.array(matchings[i_test]["iou"]).argmax()
    #             if matchings[i_test]["iou"][iou_idx] > threshold:
    #                 matched_face = matchings[i_test]['matching'][iou_idx]
    #                 predictions[-1] = prediction_dict[matched_face]['prediction_baseline_0']

    return np.array(gt_names), np.array(predictions)


def predict_bbt(train_dataset, evaluation_path="./dataset/evaluation_dict.json", prediction_path="./dataset/evaluation_dict_baseline_0.json", test_dataset=None, face_embeddings=None, mode="baseline1"):
    # train_loader = DataLoader(train_dataset, batch_size=1, shuffle=False, num_workers=0)
    # test_loader = DataLoader(test_dataset, batch_size=1, shuffle=False, num_workers=0)
    # train_dataset_anns = load_json("./dataset/new_mix_bbt_annotations.json")
    # train_dataset_lbl_to_id = load_json("./dataset/bbt_lbl_to_id.json")

    if os.path.exists(evaluation_path):
        print("Found evaluation dict in predict_bbt, loading ...")
        evaluation_dict = load_json(evaluation_path)
    else:
        print("No evaluation dict in predict_bbt, creating ...")
        evaluation_dict = {}
        for train_idx, train_ann in tqdm(train_dataset.anns.items()):
            if train_ann['face'] in evaluation_dict:
                evaluation_dict[train_ann['face']]["train_index"].append(train_idx)
                evaluation_dict[train_ann['face']]["weak_id"].append(train_dataset.lbl_to_id[train_ann['name']])
                evaluation_dict[train_ann['face']]["weak_name"].append(train_ann["name"])
            else:
                evaluation_dict[train_ann['face']] = {
                    "train_index": [train_idx],
                    "img": train_ann["img"],
                    "clip": train_ann["clip"],
                    "series": train_ann["series"],
                    "face": train_ann["face"],
                    "weak_id": [train_dataset.lbl_to_id[train_ann['name']]],
                    "weak_name": [train_ann["name"]],
                    "bbox": train_ann["bbox"],
                    "face_points": train_ann["face_points"],
                    "subtitle": train_ann["subtitle"],
                }
        save_json(evaluation_dict, evaluation_path)
        evaluation_dict = load_json(evaluation_path)

    if os.path.exists(prediction_path):
        print("Found prediction_path dict in predict_bbt, loading ...")
        prediction_dict = load_json(prediction_path)
    else:
        with torch.no_grad():
            if mode == "baseline0":
                for ann in tqdm(evaluation_dict.values()):
                    # if only one weak_name, then the choice is obvious
                    if ann['weak_name']:
                        choice = random.choice(ann['weak_name'])
                        ann["prediction_baseline_0"] = choice
                    else:
                        ann["prediction_baseline_0"] = "Unknown"
                save_json(evaluation_dict, prediction_path)
            prediction_dict = load_json(prediction_path)

    return prediction_dict

    '''
        elif mode == "baseline1":
            face_emb = []
            for train_data, face_embedding in tqdm(zip(train_loader, face_embeddings)):
                faces.append(train_data['face'])
                weak_ids.append(train_data['weak_label'])
                face_emb.append(face_embedding)
            # correct_target_names.extend(test_data['correct_target_name'])
            # correct_target_ids.append(test_data['correct_target_id'])
            # cleansed_ids.append(train_dataset.lbl_to_id[train_data['cleansed'][0]])

    # flatten list of lists
    # weak_ids = [item.item() for sublist in weak_ids for item in sublist]
    # correct_target_ids = [item.item() for sublist in correct_target_ids for item in sublist]

    if mode == "baseline0":
        results = pd.DataFrame({'faces': faces,
                                'correct_target_name': correct_target_names,
                                'correct_target_id': correct_target_ids,
                                'random_weak_label_id': weak_ids,
                                'random_weak_label_name': weak_names,
                                })
    elif mode == "baseline1":
        results = pd.DataFrame({'correct_target_name': np.nan,
                                'correct_target_id': np.nan,
                                'weak_label_ids': weak_ids,
                                'max_cluster_prediction': np.nan,
                                'cleansed': np.nan,
                                'subtitle_prediction': np.nan,
                                'face_embedding': face_emb,
                                '0': np.nan,
                                '1': np.nan,
                                '2': np.nan,
                                '3': np.nan,
                                '4': np.nan,
                                '5': np.nan,
                                '6': np.nan,
                                'min_distance': np.nan,
                                'closest_cluster': np.nan,
                                })

    return results
'''


# if __name__ == "__main__":
#     parser = argparse.ArgumentParser()
#     parser.add_argument('--episode', type=str, help='The episode number')
#     ARGS, unparsed = parser.parse_known_args()
#
#     dataset = TVQADataset(episode=ARGS.episode, split="all", transform=transforms.Compose([transforms.ToTensor()]))
#     make_excel(model=None,
#                dataset=dataset,
#                source="dataloader",
#                data_path="./dataset/frames_hq/friends_frames/",
#                path_to_faces="./dataset/friends_frames/",
#                file_name=f"s01e{ARGS.episode}.xlsx")


def visualize_cluster_distances(results):
    from sklearn.metrics.pairwise import cosine_similarity, euclidean_distances

    sorted_results = results.sort_values("cleansed") # sort based on the cleansed label
    embeddings = torch.empty(size=(3386, 512))
    for i, emb in enumerate(sorted_results["face_embedding"]):
        embeddings[i] = emb
    dists = euclidean_distances(embeddings)

    plt.imshow(dists, cmap='plasma', interpolation='nearest')
    # plt.xticks(range(len(sorted_results["cleansed"].values)), sorted_results["cleansed"].values, rotation=90, fontsize=6)
    # plt.yticks(range(len(sorted_results["cleansed"].values), sorted_results["cleansed"].values, rotation=0, fontsize=6)

    cb = plt.colorbar()
    cb.set_label("distances")
    plt.title("distances of points of each cluster to other cluster")

    plt.savefig('small__distances.pdf')
    plt.clf()


def convert_dfoftensors_to_tensor(df):
    new_tensor = torch.empty(size=(len(df), len(df.iloc[0])))
    for i, item in enumerate(df):
        new_tensor[i] = item
    return new_tensor


def calc_distances(train_dataset, results, num_classes, alpha=None):
    unknown_id = train_dataset.lbl_to_id["Unknown"]
    from scipy.spatial import distance
    for i, unknown in tqdm(results[results["cleansed"] == unknown_id].iterrows()):
        for cluster in range(num_classes):
            unknown_cluster_emb = unknown["face_embedding"]
            other_cluster_embs = results.loc[results["cleansed"] == cluster, "face_embedding"]
            if not other_cluster_embs.empty: #if this is not an empty series
                other_cluster_embs_torch = convert_dfoftensors_to_tensor(other_cluster_embs)
                c_dists = distance.cdist(unknown_cluster_emb[None, :], other_cluster_embs_torch, 'euclidean')
            else:
                # print(f"cluster {cluster} is empty!")
                c_dists = np.array([np.inf])
            if alpha is not None:
                if cluster == unknown_id:
                    results.loc[i, (f'{cluster}')] = c_dists.mean() * alpha
                else:
                    results.loc[i, (f'{cluster}')] = c_dists.mean() * (1-alpha)
            else:
                results.loc[i, (f'{cluster}')] = c_dists.mean()
    results["min_distance"] = results[[str(x) for x in range(num_classes)]].min(axis=1)
    results["closest_cluster"] = results[[str(x) for x in range(num_classes)]].idxmin(axis=1)
    results.loc[results["cleansed"] != unknown_id, "closest_cluster"] = results.loc[results["cleansed"] != unknown_id, "cleansed"]
    results["closest_cluster"] = results["closest_cluster"].astype(np.int64)
    return results


def calc_distances_with_prototypes(lbl_to_id, results, num_classes, alpha=None):
    # unknown_id = lbl_to_id["Unknown"]
    unknown_id = lbl_to_id["Sheldon"]

    from scipy.spatial import distance
    prototype_embeddings = {}
    for cluster in tqdm(range(num_classes)):
        cluster_embs = results.loc[results["cleansed"] == cluster, "face_embedding"]
        if cluster_embs.empty:
            prototype_embeddings[cluster] = np.ones((1, 512)) * np.inf
        else:
            cluster_embs = convert_dfoftensors_to_tensor(cluster_embs)
            prototype_embeddings[cluster] = cluster_embs.mean(axis=0, keepdim=True)

    unknown_cluster_embs = convert_dfoftensors_to_tensor(results.loc[results["cleansed"] == unknown_id, "face_embedding"])
    for prototype_idx, prototype_emb in tqdm(prototype_embeddings.items()):
        c_dists = distance.cdist(unknown_cluster_embs, prototype_emb, 'euclidean')
        if alpha is not None:
            if prototype_idx == unknown_id:
                c_dists = c_dists * alpha
            else:
                c_dists = c_dists * (1 - alpha)
        results.loc[results["cleansed"] == unknown_id, (f'{prototype_idx}')] = c_dists

    results["min_distance"] = results[[str(x) for x in range(num_classes)]].min(axis=1)
    results["closest_cluster"] = results[[str(x) for x in range(num_classes)]].idxmin(axis=1)
    results.loc[results["cleansed"] != unknown_id, "closest_cluster"] = results.loc[results["cleansed"] != unknown_id, "cleansed"]
    results["closest_cluster"] = results["closest_cluster"].astype(np.int64)
    return results


def recluster_unknowns(results, num_classes=7, num_clusters=7):
    from scipy.spatial import distance
    results['recluster_unk'] = np.nan
    results['hac7_id'] = np.nan
    embeddings = convert_dfoftensors_to_tensor(results.loc[results["cleansed"] == 6, 'face_embedding'])
    hac7_id = AgglomerativeClustering(n_clusters=num_clusters).fit_predict(embeddings.numpy())
    results.loc[results["cleansed"] == 6, 'hac7_id'] = hac7_id
    unknowns = results[results["cleansed"] == 6]

    for hacid in range(num_clusters):
        for cluster in range(num_clusters):
            new_cluster_embs = unknowns.loc[unknowns["hac7_id"] == hacid, 'face_embedding']
            other_cluster_embs = results.loc[results["cleansed"] == cluster, 'face_embedding']

            new_cluster_embs_torch = convert_dfoftensors_to_tensor(new_cluster_embs)
            other_cluster_embs_torch = convert_dfoftensors_to_tensor(other_cluster_embs)
            c_dists = distance.cdist(new_cluster_embs_torch, other_cluster_embs_torch, 'euclidean')
            results.loc[results["hac7_id"] == hacid, f'{cluster}'] = c_dists.mean()

    results["min_distance"] = results[[str(x) for x in range(num_classes)]].min(axis=1)
    results["closest_cluster"] = results[[str(x) for x in range(num_classes)]].idxmin(axis=1)
    results.loc[results["cleansed"] != 6, "closest_cluster"] = results.loc[results["cleansed"] != 6, "cleansed"]
    results["closest_cluster"] = results["closest_cluster"].astype(np.int64)
    return results


def calc_iou(bbox1, bbox2):
    x_left = max(bbox1[0], bbox2[0])
    y_top = max(bbox1[1], bbox2[1])
    x_right = min(bbox1[2], bbox2[2])
    y_bottom = min(bbox1[3], bbox2[3])

    # no intersection
    if x_right < x_left or y_bottom < y_top:
        return 0.0
    intersection_area = (x_right - x_left) * (y_bottom - y_top)
    bb1_area = (bbox1[2] - bbox1[0]) * (bbox1[3] - bbox1[1])
    bb2_area = (bbox2[2] - bbox2[0]) * (bbox2[3] - bbox2[1])

    iou = intersection_area / float(bb1_area + bb2_area - intersection_area)
    return iou

def visualize(cfg):
    seed = 10
    np.random.seed(seed)
    model_mode = "facenet_pretrained"
    # model_mode = "facenet_reclassified"
    # model_mode = "nothing"
    data_mode = "test"
    if cfg.TRAINING.series == "bbt":
        num_classes = 9
    elif cfg.TRAINING.series == "friends":
        num_classes = 7
    num_clusters = 30
    exp = cfg.TRAINING.project_dir.split('/')[-2]
    print(f"this is experiment {exp}")
    epoch = 99
    print(f"cuda is available: {torch.cuda.is_available()}")
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

    vis_path = f"./output/visualize/{exp}"
    Path(vis_path).mkdir(parents=True, exist_ok=True)
    logging.basicConfig(filename=f'{vis_path}/log.out', filemode='a', level=logging.INFO,
                        format='%(asctime)s %(name)s - %(levelname)s - %(message)s')
    logger = logging.getLogger('visualize')

    if model_mode == "resnet_pretrained":
        model = models.ResNet(block=BasicBlock, layers=[2, 2, 2, 2], num_classes=num_classes)
        # loading model checkpoint
        checkpoint = torch.load(f"./output/{exp}/model/epoch_{epoch}.tar", map_location=torch.device(device))
        model.load_state_dict(checkpoint['model_state_dict'])
        model.eval()
    elif model_mode == "facenet_pretrained":
        model = InceptionResnetV1(pretrained='vggface2').eval()
    elif model_mode == "facenet_reclassified":
        model = VGGSupervised(cfg, num_classes)
        # model = VGGFacePlus(cfg, num_classes)
        checkpoint = torch.load(f"./output/{exp}/model/epoch_{epoch}.tar", map_location=torch.device(device))
        model.load_state_dict(checkpoint['model_state_dict'])
        model.eval()
    elif model_mode == "facenet_reclassified_VGGFaceSubtitle":
        model = VGGFaceSubtitle(cfg, num_classes)
        checkpoint = torch.load(f"./output/{exp}/model/epoch_{epoch}.tar", map_location=torch.device(device))
        model.load_state_dict(checkpoint['model_state_dict'])
        model.eval()

    if exp.startswith("oracle_bbt") or exp == 'multilabel_bbt':
        logging.info("bbt oracle...")
        test_dataset = TVQADataset(series=cfg.TRAINING.series, split="test", transform=get_test_transforms(series=cfg.TRAINING.series))
        results = evaluate(model, test_dataset, model_mode)
        prediction_dict = load_json(f"./dataset/bbt_9_10.json")
        prediction_dict = save_predictions(test_dataset.id_to_lbs, results,
                                           f"./dataset/bbt_9_10_multilabel_prediction.json",
                                           prediction_dict, prediction_mode='model_prediction')
        # prediction_dict = load_json("./dataset/bbt_9_10_oracle_prediction.json")
        matchings = load_json("./dataset/box_matchings_with_iou.json")
        gt_names, predictions_names = calculate_bbt(matchings, prediction_dict, test_dataset, prediction_mode='model_prediction')
        logging.info(f"number of predictions is {predictions_names.shape, gt_names.shape}")
        logging.info(f"per sample accuracy of model_prediction in oracle is {calc_accuracies_bbt(gt_names, predictions_names)}")
        accuracies = calc_per_class_accuracy_bbt(test_dataset, gt_names, predictions_names)
        logging.info(f"mean per class accuracy with unknown: {accuracies[0].mean()}")
        # logging.info(f"mean per class accuracy without unknown: {accuracies[0][:-1].mean()}")
        logging.info(f"per class accuracies: {accuracies}")
        logging.info(f"per class precision and recalls: {calc_per_class_prec_recall_bbt(test_dataset, gt_names, predictions_names)}")
        print("koko")
        # prediction_mode = 'model_prediction'
        # print(
        #     f"per sample accuracy of model_prediction in {data_mode} dataset and model {model_mode} is {calc_accuracies(results, mode=prediction_mode)}")
        # accuracies = calc_per_class_accuracy(test_dataset, results, mode=prediction_mode)
        # print(f"mean per class accuracy: {accuracies[0].mean()}")
        # print(f"per class accuracies: {accuracies}")
        # print(f"per class precision and recalls: {calc_per_class_prec_recall(results, mode=prediction_mode)}")
    elif exp == "bbt_baseline0":
        logging.info("bbt baseline 0...")
        train_dataset = TVQADataset(series=cfg.TRAINING.series, split="train", transform=get_test_transforms(series=cfg.TRAINING.series))
        test_dataset = TVQADataset(series=cfg.TRAINING.series, split="test", transform=get_test_transforms(series=cfg.TRAINING.series))
        prediction_dict = predict_bbt(train_dataset, mode="baseline0")
        matchings = match_bbox_2(test_dataset, prediction_dict)
        gt_names, predictions_names = calculate_bbt(matchings, prediction_dict, test_dataset)

        logging.info(f"per sample accuracy of model_prediction in baseline 0 is {calc_accuracies_bbt(gt_names, predictions_names)}")
        accuracies = calc_per_class_accuracy_bbt(train_dataset, gt_names, predictions_names)
        logging.info(f"mean per class accuracy with unknown: {accuracies[0].mean()}")
        logging.info(f"mean per class accuracy without unknown: {accuracies[0][:-1].mean()}")
        logging.info(f"per class accuracies: {accuracies}")
        logging.info(f"per class precision and recalls: {calc_per_class_prec_recall_bbt(train_dataset, gt_names, predictions_names)}")
    elif exp == "iou_matchings":
        logging.info("iou_matchings...")
        train_dataset = TVQADataset(series=cfg.TRAINING.series, split="train", transform=get_test_transforms(series=cfg.TRAINING.series))
        test_dataset = TVQADataset(series=cfg.TRAINING.series, split="test", transform=get_test_transforms(series=cfg.TRAINING.series))
        prediction_dict = predict_bbt(train_dataset, mode="baseline0")
        matchings = match_bbox_2(test_dataset, prediction_dict)
    elif exp == "calculate_embeddings":
        logging.info("calculating embeddings...")
        train_dataset = TVQADataset(series=cfg.TRAINING.series, split="train", transform=get_test_transforms(series=cfg.TRAINING.series))
        face_embeddings = calculate_embeddings(model, emb_path=f"./output/evaluate_bbt/model/new_bbt_face_embeddings.pt",
                                               dataset=train_dataset, model_mode=model_mode, preload=False)
    elif exp == "ablation_num_clusters_friends":
        train_dataset = TVQADataset(series=cfg.TRAINING.series, split="train", transform=get_test_transforms(series=cfg.TRAINING.series))
        test_dataset = TVQADataset(series=cfg.TRAINING.series, split="test", transform=get_test_transforms(series=cfg.TRAINING.series))
        face_embeddings = calculate_embeddings(model, emb_path=f"./dataset/bigger_bb/face_embeddings_friends.pt",dataset=train_dataset, model_mode=model_mode, preload=True)
        results = evaluate_self_supervised(train_dataset, test_dataset, face_embeddings, mode="baseline1")
        exp_num_clusters(range(7, 31), results, face_embeddings, test_dataset)
    elif exp == "clustering_and_cleansing_and_closest_kmeans" or exp=="test" or exp=="cleansing_sheldon":
        logging.info("clustering_and_cleansing ...")
        train_dataset = TVQADataset(series=cfg.TRAINING.series, split="train", transform=get_test_transforms(series=cfg.TRAINING.series))
        test_dataset = TVQADataset(series=cfg.TRAINING.series, split="test", transform=get_test_transforms(series=cfg.TRAINING.series))
        face_embeddings = calculate_embeddings(model, emb_path=f"./dataset/bbt_face_embeddings.pt", dataset=train_dataset, model_mode=model_mode, preload=True)
        results = prepare_result(train_dataset, face_embeddings)
        prediction_mode = "max_cluster_prediction"
        results, clustering_ids = predict_with_clustering(results, face_embeddings, n_clusters=num_clusters, knn_graph=None, pred_mode=prediction_mode)
        # cleanse_labels(train_dataset, results, file_path=f"./dataset/prediction.json")
        results = calc_distances_with_prototypes(train_dataset.lbl_to_id, results, num_classes)
        # results = calc_distances(train_dataset, results, num_classes, alpha=None)
        prediction_dict = load_json(f"./dataset/bbt_train_annotations.json")

        prediction_dict = save_predictions(train_dataset.id_to_lbs, results, f"./dataset/bbt_train_annotations_mediaeval.json", prediction_dict, prediction_mode='M2')


        prediction_dict = save_predictions(train_dataset.id_to_lbs, results, f"./dataset/bbt_train_annotations_sheldon.json", prediction_dict, prediction_mode='closest_cluster')
        matchings = load_json("./dataset/box_matchings_with_iou.json")
        gt_names, predictions_names = calculate_bbt(matchings, prediction_dict, test_dataset, prediction_mode='closest_cluster')

        logging.info(f"per sample accuracy of model_prediction in baseline 0 is {calc_accuracies_bbt(gt_names, predictions_names)}")
        accuracies = calc_per_class_accuracy_bbt(train_dataset, gt_names, predictions_names)
        logging.info(f"mean per class accuracy with unknown: {accuracies[0].mean()}")
        logging.info(f"mean per class accuracy without unknown: {accuracies[0][:-1].mean()}")
        logging.info(f"per class accuracies: {accuracies}")
        logging.info(f"per class precision and recalls: {calc_per_class_prec_recall_bbt(train_dataset, gt_names, predictions_names)}")

        gt_names, predictions_names = calculate_bbt(matchings, prediction_dict, test_dataset, prediction_mode='cleansed')
        logging.info(
            f"per sample accuracy of model_prediction in baseline 0 is {calc_accuracies_bbt(gt_names, predictions_names)}")
        accuracies = calc_per_class_accuracy_bbt(train_dataset, gt_names, predictions_names)
        logging.info(f"mean per class accuracy with unknown: {accuracies[0].mean()}")
        logging.info(f"mean per class accuracy without unknown: {accuracies[0][:-1].mean()}")
        logging.info(f"per class accuracies: {accuracies}")
        logging.info(
            f"per class precision and recalls: {calc_per_class_prec_recall_bbt(train_dataset, gt_names, predictions_names)}")
    elif exp == "ours_s09_s10":
        logging.info("ours_s09_s10 ...")
        train_dataset = TVQADataset(series=cfg.TRAINING.series, split="train", transform=get_test_transforms(series=cfg.TRAINING.series))
        test_dataset = TVQADataset(series=cfg.TRAINING.series, split="test", transform=get_test_transforms(series=cfg.TRAINING.series))

        prediction_dict = load_json(f"./dataset/bbt_train_annotations_faster_closestcluster.json")
        matchings = load_json("./dataset/box_matchings_with_iou.json")
        gt_names, predictions_names = calculate_bbt(matchings, prediction_dict, test_dataset, prediction_mode='closest_cluster')

        logging.info(
            f"per sample accuracy of model_prediction in baseline 0 is {calc_accuracies_bbt(gt_names, predictions_names)}")
        accuracies = calc_per_class_accuracy_bbt(train_dataset, gt_names, predictions_names)
        logging.info(f"mean per class accuracy with unknown: {accuracies[0].mean()}")
        logging.info(f"per class accuracies: {accuracies}")
        logging.info(
            f"per class precision and recalls: {calc_per_class_prec_recall_bbt(train_dataset, gt_names, predictions_names)}")
        print("booboo")
    elif exp == "friends_prediction":
        train_dataset = TVQADataset(series="friends", split="train", transform=get_test_transforms(series="friends"))
        test_dataset = TVQADataset(series="friends", split="test", transform=get_test_transforms(series="friends"))
        # face_embeddings = calculate_embeddings(model, emb_path=f"./dataset/bigger_bb/face_embeddings_friends.pt",
        #                                        dataset=train_dataset, model_mode=model_mode, preload=True)
        face_embeddings = calculate_embeddings(model, emb_path=f"./output/distances/model/face_embeddings.pt",dataset=train_dataset, model_mode=model_mode, preload=True)

        results = evaluate_self_supervised(train_dataset, test_dataset, face_embeddings, mode="baseline1")
        # results = calc_distances(train_dataset, results, num_classes, alpha=None)

        # prediction_dict = load_json(f"./dataset/train_annotations.json")
        # test_dict = load_json(f"./dataset/test_annotations.json")
        # prediction_dict = save_predictions(train_dataset.id_to_lbs, results, f"./dataset/friends_prediction_dict.json",
        #                                    prediction_dict, prediction_mode='closest_cluster')
        #
        # for ann, ann_test in zip(prediction_dict.values(), test_dict.values()):
        #     ann["correct_target_name"] = ann_test["target_name"]
        #     ann["correct_target_id"] = train_dataset.lbl_to_id[ann_test["target_name"]]
        # save_json(prediction_dict, file_path=f"./dataset/friends_prediction_dict.json")

        results, clustering_ids = predict_with_clustering(results, face_embeddings, n_clusters=num_clusters, knn_graph=None, pred_mode="max_cluster_prediction")
        prediction_mode = "M2"
        print(
            f"per sample accuracy of model_prediction in {data_mode} dataset and model {model_mode} is {calc_accuracies(results, mode=prediction_mode)}")
        accuracies = calc_per_class_accuracy(test_dataset, results, mode=prediction_mode)
        print(f"mean per class accuracy: {accuracies[0].mean()}")
        print(f"per class accuracies: {accuracies}")
        print(f"per class precision and recalls: {calc_per_class_prec_recall(results, mode=prediction_mode)}")
        visualize_cluster_distances(results)
    # result_matching = match_bbox(train_dataset, test_dataset)

    # results = evaluate(model, test_dataset, model_mode)
    # face_embeddings = calculate_embeddings(model, emb_path=f"./output/evaluate_bbt/model/bbt_face_embeddings.pt", dataset=train_dataset, model_mode=model_mode, preload=True)
    # face_embeddings = calculate_embeddings(model, emb_path=f"./output/evaluate_bbt/model/face_embeddings.pt",
    #                                        dataset=train_dataset, model_mode=model_mode, preload=True)
    # face_embeddings_train = calculate_embeddings(model, emb_path=f"./dataset/bigger_bb/face_embeddings_train_{cfg.TRAINING.series}.pt",
    #                                        dataset=train_dataset, model_mode="facenet_pretrained", preload=True)
    # face_embeddings_test = calculate_embeddings(model, emb_path=f"./dataset/bigger_bb/face_embeddings_test_{cfg.TRAINING.series}.pt",
    #                                        dataset=test_dataset, model_mode="facenet_pretrained", preload=True)
    # tsne_grid = TSNE(random_state=seed, n_iter=2000).fit_transform(face_embeddings_train.detach().numpy())
    # torch.save(tsne_grid, "./output/evaluate_bbt/model/tsne_grid.pt")
    # tsne_grid = torch.load("./output/evaluate_bbt/model/tsne_grid.pt")
#     a == tsne_grid
#     predict_bbt(train_dataset, mode="baseline0")

    # gt_names, predictions_names = match_bbox_2(test_dataset, threshold=0)
#     print(f"per sample accuracy of model_prediction in baseline 0 is {calc_accuracies_bbt(gt_names, predictions_names)}")
#     accuracies = calc_per_class_accuracy_bbt(train_dataset, gt_names, predictions_names)
#     print(f"mean per class accuracy with unknown: {accuracies[0].mean()}")
#     print(f"mean per class accuracy without unknown: {accuracies[0][:-1].mean()}")
#     print(f"per class accuracies: {accuracies}")
#     print(f"per class precision and recalls: {calc_per_class_prec_recall_bbt(train_dataset, gt_names, predictions_names)}")

    # results.to_pickle("./output/evaluate_bbt_baseline0/results_baseline0.pkl")
    # results = pd.read_pickle("./output/evaluate_bbt_baseline0/results_baseline0.pkl")
    # results = evaluate_self_supervised(train_dataset, test_dataset, face_embeddings_all)

    # results = evaluate_oracle_supervised(model, test_dataset, face_embeddings)
# #
# #     # knn_graph = build_graph(face_embeddings, train_dataset, neighbours=3385, no_edge=True)
# #     num_clusters = 10
#     results = prepare_result(train_dataset, face_embeddings)
#     prediction_mode = "max_cluster_prediction"
#     results, clustering_ids = predict_with_clustering(results, face_embeddings, n_clusters=num_clusters, knn_graph=None, pred_mode=prediction_mode)
#     cleanse_labels(train_dataset, results, file_path=f"./dataset/cleansed_bbt_annotations_{global_cfg.TRAINING.clustering}_{global_cfg.TRAINING.kmeans_batch_size}.json")
#
#     # visualize_cluster_distances(results)


    # results = evaluate_ep_1_5(train_dataset, face_embeddings_all)
    # alpha = 1.0
    # results = calc_distances(results, num_classes, alpha=alpha)




    # results = recluster_unknowns(results, num_classes, num_clusters=7)
#     # results, plt, fig = clustering_with_gmm(results, face_embeddings, tsne_grid)
#     # fig.savefig(os.path.join(f"{vis_path}", f"lala.pdf"))
#     # plt.clf()
#     # exp_num_clusters(range(7, 20), results, embeddings)
#     # results, clustering_ids = predict_with_clustering(results, face_embeddings, n_clusters=num_clusters, pred_mode="max_cluster_prediction")
#     # results, sub_clustering_ids = predict_with_clustering(results, subtitle_embeddings, n_clusters=num_clusters, pred_mode="subtitle_prediction")
#
#
    # prediction_mode = 'model_prediction'
    # print(f"per sample accuracy of model_prediction in {data_mode} dataset and model {model_mode} is {calc_accuracies(results, mode=prediction_mode)}")
    # accuracies = calc_per_class_accuracy(test_dataset, results, mode=prediction_mode)
    # print(f"mean per class accuracy: {accuracies[0].mean()}")
    # print(f"per class accuracies: {accuracies}")
    # print(f"per class precision and recalls: {calc_per_class_prec_recall(results, mode=prediction_mode)}")
#
#     fig, plt = visualize_tsne(tsne_grid, results["closest_cluster"].values.tolist(), train_dataset.id_to_lbs)
#     fig.savefig(os.path.join(f"{vis_path}", f"{data_mode}_{model_mode}_closest_cluster.pdf"))
#     plt.clf()
#
#     fig, plt = visualize_tsne(tsne_grid, results["cleansed"].values.tolist(), train_dataset.id_to_lbs)
#     fig.savefig(os.path.join(f"{vis_path}", f"{data_mode}_{model_mode}_cleansed.pdf"))
#     plt.clf()
#

    # fig, plt = visualize_tsne(tsne_grid, results.loc[len(train_dataset) - len(test_dataset):len(train_dataset)-1,"correct_target_id"].values.tolist(), train_dataset.id_to_lbs)
    # fig.savefig(os.path.join(f"{vis_path}", f"{data_mode}_{model_mode}_groundtruth.pdf"))
    # plt.clf()
    # fig, plt = visualize_tsne(tsne_grid, results.loc[len(train_dataset) - len(test_dataset):len(train_dataset)-1,"closest_cluster"].values.tolist(), train_dataset.id_to_lbs)
    # fig.savefig(os.path.join(f"{vis_path}", f"{data_mode}_{model_mode}_closest_cluster.pdf"))
    # plt.clf()
    # print(f"accuracy of model_prediction in {data_mode} dataset and model {model_mode} is"
    #       f" {calc_accuracies(results.loc[len(train_dataset) - len(test_dataset) :len(train_dataset)-1], mode='closest_cluster')}")
    # print(f"per class accuracies: {calc_per_class_accuracy(train_dataset, results.loc[len(train_dataset) - len(test_dataset) :len(train_dataset)-1], mode='closest_cluster')}")
    # print(f"per class precision and recalls: {calc_per_class_prec_recall(results.loc[len(train_dataset) - len(test_dataset):len(train_dataset)-1], mode='closest_cluster')}")
#
#     ###########################################visualize#################################################
#
#
#     # fig, plt = visualize_tsne(tsne_grid, results["weak_label"].values.tolist(), dataset.id_to_lbs)
#     # fig.savefig(os.path.join(f"{vis_path}", f"{data_mode}_{model_mode}_weaklabels.pdf"))
#     # plt.clf()
#
#     #
#     fig, plt = visualize_tsne(tsne_grid, clustering_ids, train_dataset.id_to_lbs)
#     fig.savefig(os.path.join(f"{vis_path}", f"{data_mode}_{model_mode}_hac8clusters.pdf"))
#     plt.clf()
#
#
#     # # logging.info(f"accuracy of maximum clustering in {data_mode} dataset and model {model_mode} is "
#     # #       f"{calc_accuracies(results, mode='max_cluster_prediction')}")
#     # # print(f"accuracy of maximum clustering in {data_mode} dataset and model {model_mode} is "
#     # #       f"{calc_accuracies(results, mode='max_cluster_prediction')}")
#     # # logging.info(f"per class accuracies: {calc_per_class_accuracy(results, mode='max_cluster_prediction')}")
#     # print(f"per class accuracies: {calc_per_class_accuracy(results, mode='max_cluster_prediction')}")
#     logger.info(f"accuracy of model_prediction in {data_mode} dataset and model {model_mode} is"
#                  f" {calc_accuracies(results, mode='max_cluster_prediction')}")
#     print(f"accuracy of model_prediction in {data_mode} dataset and model {model_mode} is"
#           f" {calc_accuracies(results, mode='max_cluster_prediction')}")
#     logger.info(
#         f"per class accuracies: {calc_per_class_accuracy(test_dataset, results, mode='max_cluster_prediction')}")
#     print(f"per class accuracies: {calc_per_class_accuracy(test_dataset, results, mode='max_cluster_prediction')}")
#
#     logger.info(
#         f"per class precision and recalls: {calc_per_class_prec_recall(results, mode='max_cluster_prediction')}")
#     print(f"per class precision and recalls: {calc_per_class_prec_recall(results, mode='max_cluster_prediction')}")
#
#
#     # fig = plt.figure(figsize=(8, 8))
#     # plt.style.use('seaborn-darkgrid')
#     # ax = plt.subplot(aspect='equal')
#     # plt.plot(np.array([0.1, 0.4, 0.5, 0.52, 0.55, 0.6, 0.9]), np.array(
#     #     [0.696776646489589, 0.696776646489589, 0.7631103930418455, 0.7740073757551987, 0.7277433039443212,
#     #      0.7218788892129113, 0.7218788892129113]), '-o', label='mean per class accuracy')
#     # plt.plot(np.array([0.1, 0.4, 0.5, 0.52, 0.55, 0.6, 0.9]), np.array(
#     #     [0.7409923213230951, 0.7409923213230951, 0.7359716479621973, 0.7002362669816893, 0.5593620791494389,
#     #      0.544595392793857, 0.544595392793857]), '-o', label='sample-level accuracy')
#     # plt.xlim(0.0, 1.0)
#     # plt.ylim(0.6, 0.8)
#     # plt.legend()
#     # plt.xlabel("alpha")
#     # ax.axis('tight')
#     # fig.savefig(os.path.join(f"{vis_path}", f"accuracy_alphas.pdf"))
#     # plt.clf()
# #########################
#     # fig = plt.figure(figsize=(8, 8))
#     # plt.style.use('seaborn-darkgrid')
#     # ax = plt.subplot(aspect='equal')
#     # plt.plot(np.arange(7), np.array([0.54754098, 0.67146283, 0.82484725, 0.71489362, 0.64166667,0.63055556, 0.84646962]), '-o', label='0.1')
#     # plt.plot(np.arange(7), np.array([0.54754098, 0.67146283, 0.82484725, 0.71489362, 0.64166667,0.63055556, 0.84646962]), '-o', label='0.4')
#     # plt.plot(np.arange(7), np.array([0.7147541 , 0.83453237, 0.91242363, 0.80425532, 0.73888889,0.70555556, 0.63136289]), '-o', label='0.5')
#     # plt.plot(np.arange(7), np.array([0.79672131, 0.86091127, 0.95315682, 0.85531915, 0.78055556,0.70833333, 0.46305419]), '-o', label='0.52')
#     # plt.plot(np.arange(7), np.array([0.81311475, 0.882494  , 0.96741344, 0.85957447, 0.82222222,0.70833333, 0.0410509 ]), '-o', label='0.55')
#     # plt.plot(np.arange(7), np.array([0.81311475, 0.882494  , 0.96741344, 0.85957447, 0.82222222,0.70833333, 0.        ]), '-o', label='0.6')
#     # plt.plot(np.arange(7), np.array([0.81311475, 0.882494  , 0.96741344, 0.85957447, 0.82222222,0.70833333, 0.        ]), '-o', label='0.9')
#     # plt.xlim(0, 6)
#     # plt.ylim(0.2, 1)
#     # plt.legend()
#     # ax.axis('tight')
#     # fig.savefig(os.path.join(f"{vis_path}", f"alphas.pdf"))
#     # plt.clf()
# #############
#     # fig = plt.figure(figsize=(8, 8))
#     # plt.style.use('seaborn-darkgrid')
#     # ax = plt.subplot(aspect='equal')
#     # plt.plot(np.arange(4), np.array([0, 0.6967766471428571, 0.7631103942857143, 0.7740073757551987]), '-o',label='mean per class accuracy')
#     # plt.plot(np.arange(4), np.array([0, 0.7409923213230951, 0.7359716479621973, 0.7002DataLoader362669816893]), '-o',label='sample accuracy')
#     # plt.xlim(0, 4)
#     # plt.ylim(0.6, 0.8)
#     # plt.legend()
#     # plt.xticks(range(4), np.array(["random from weaklabel", "max cluster", "closest cluster", "alpha 0.52"]), rotation=45, fontsize=6)
#     # plt.xlabel("methods")
#     # ax.axis('tight')
#     # fig.savefig(os.path.join(f"{vis_path}", f"accuracy_methods.pdf"))
#     # plt.clf()
# ##############
#     # fig = plt.figure(figsize=(8, 8))
#     # plt.style.use('seaborn-darkgrid')
#     # ax = plt.subplot(aspect='equal')
#     # plt.plot(np.arange(7), np.array([0.17704918, 0.17026379, 0.27494908, 0.14468085, 0.2       ,0.19166667, 0.62068966]), '-o', label='random from weaklabel')
#     # plt.plot(np.arange(7), np.array([0.54754098, 0.67146283, 0.82484725, 0.71489362, 0.64166667, 0.63055556, 0.84646962]), '-o', label='max cluster')
#     # plt.plot(np.arange(7), np.array([0.7147541 , 0.83453237, 0.91242363, 0.80425532, 0.73888889, 0.70555556, 0.63136289]), '-o', label='closest cluster')
#     # plt.plot(np.arange(7), np.array([0.79672131, 0.86091127, 0.95315682, 0.85531915, 0.78055556, 0.70833333, 0.46305419]), '-o', label='alpha 0.52')
#     # plt.xlim(0, 6)
#     # plt.ylim(0.2, 1)
#     # plt.legend()
#     # ax.axis('tight')
#     # fig.savefig(os.path.join(f"{vis_path}", f"per_class_accuracies_methods.pdf"))
#     # plt.clf()


def self_supervised_train_test_split():
    test = load_json("./dataset/self_supervised_dataset/3_test_annotations.json")
    dev = load_json("./dataset/self_supervised_dataset/3_dev_annotations.json")
    all = {}
    i = 0
    while i < len(test.keys()) + len(dev.keys()):
        for ann in test.values():
            all[i] = ann
            i += 1
        for ann in dev.values():
            all[i] = ann
            i += 1
    print(len(test),len(dev), len(all))
    save_json(all, "./dataset/self_supervised_dataset/all_annotations.json")
    train = {}
    test = {}
    for i, ann in all.items():
        train[i] = {key: ann[key] for key in ann.keys() if key in ['face', 'name', 'img', 'subtitle', 'clip', 'series']}
        test[i] = {key: ann[key] for key in ann.keys() if key in ['face', 'img', 'subtitle', 'clip', 'series', 'target_name']}
    save_json(train, "./dataset/self_supervised_dataset/train_annotations.json")
    save_json(test, "./dataset/self_supervised_dataset/test_annotations.json")



if __name__ == "__main__":
    # config priority:
    # 1. arguments
    # 2. config.yaml file
    # 3. defaults.py file
    args = default_argument_parser().parse_args()
    print("Command Line Args:", args)
    cfg = get_cfg()
    cfg.merge_from_file(args.config_file)
    cfg.merge_from_list(args.opts)
    cfg.freeze()

    print("Command line arguments: " + str(args))
    print("Running with full config:\n{}".format(cfg))
    Path(cfg.TRAINING.project_dir).mkdir(parents=True, exist_ok=True)
    Path(os.path.join(cfg.TRAINING.project_dir, "model")).mkdir(parents=True, exist_ok=True)
    config_path = os.path.join(cfg.TRAINING.project_dir, "config.yaml")
    with PathManager.open(config_path, "w") as f:
        f.write(cfg.dump())
    print("Full config saved to {}".format(os.path.abspath(config_path)))
    set_global_cfg(cfg)
    visualize(cfg)

def qualitative_results():
    import xlsxwriter
    import os
    from tvqa_dataset import save_json, load_json
    import numpy as np
    prediction_dict = load_json("./dataset/bbt_train_annotations_faster_closestcluster.json")
    img_to_faces = {}
    for i, ann in prediction_dict.items():
        img = ann['clip']+'_'+ann['img']
        if img in img_to_faces:
            img_to_faces[img]['train_id'].append(i)
            img_to_faces[img]['face'].append(ann['face'])
        else:
            img_to_faces[img] = {'train_id': [i], 'face': [ann['face']]}



    face_path = "./dataset/bbt_frames/"
    img_path = "./dataset/frames_hq/bbt_frames/"
    workbook = xlsxwriter.Workbook(f"./dataset/friends_qualitative_results.xlsx")
    worksheet = workbook.add_worksheet()
    worksheet.write('A1', 'img')
    worksheet.write('B1', 'subtitle')
    worksheet.write('C1', 'faces')
    worksheet.write('D1', 'stage1')
    worksheet.write('E1', 'stage2')
    worksheet.write('F1', 'stage3')
    worksheet.write('G1', 'face1')
    worksheet.write('H1', 'face2')
    worksheet.write('I1', 'face3')
    worksheet.write('J1', 'face4')

    idx = 1
    for img, values in img_to_faces.items():
        train_ids = values['train_id']
        faces = values['face']
        if len(faces) != 3:
            continue
        ann = prediction_dict[train_ids[0]]
        # if len(ann['name']) < 5:
        #     continue
        image = "_".join(ann['face'].split('_')[:-1])
        if image not in ['s03e01_seg02_clip_15_00032', 's01e13_seg02_clip_11_00170','s04e06_seg02_clip_16_00018', 's04e11_seg02_clip_02_00057','s04e11_seg02_clip_02_00061','s04e13_seg02_clip_00_00076','s04e15_seg02_clip_01_00147','s05e08_seg01_clip_00_00099','s05e10_seg02_clip_04_00038','s06e23_seg02_clip_07_00135','s07e11_seg02_clip_18_00029','s08e03_seg02_clip_05_00082','s08e10_seg02_clip_08_00102','s08e23_seg02_clip_04_00057','s09e17_seg02_clip_08_00060','s09e22_seg02_clip_07_00075','s09e24_seg02_clip_09_00071','s09e24_seg02_clip_12_00055','s10e06_seg02_clip_12_00123','s03e21_seg02_clip_13_00094','s10e01_seg02_clip_13_00019','s04e04_seg02_clip_04_00044','s03e01_seg02_clip_15_00051','s04e24_seg01_clip_01_00225']:
            continue
        worksheet.insert_image(idx, 0, os.path.join(img_path+ann['clip'], ann['img']+'.jpg'))
        worksheet.write(idx, 1, ann['subtitle'])
        worksheet.write(idx, 2, " ".join(faces))
        stage1, stage2, stage3 = ann['name'], [], []
        for i, (train_id, face) in enumerate(zip(train_ids, faces)):
            ann = prediction_dict[train_id]
            stage2.append(ann['cleansed'])
            stage3.append(ann['closest_cluster'])
            worksheet.insert_image(idx, i+6, os.path.join(face_path, face))
        # if set(stage2) == set(stage3):
        #     continue
        print(f"image: {ann['img']+'.jpg'}, subtitle: {ann['subtitle']}, faces: {faces}, stage1: {stage1}, stage2: {stage2}, stage3: {stage3}")
        worksheet.write(idx, 3, " ".join(stage1))
        worksheet.write(idx, 4, " ".join(stage2))
        worksheet.write(idx, 5, " ".join(stage3))
        idx += 1
    workbook.close()